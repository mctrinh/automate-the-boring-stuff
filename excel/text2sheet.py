"""Insert the contents of multiple text files into a single spreadsheet."""

import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def get_input_path():
    """Get the path to the folder containing the text files."""
    return input('Enter the absolute path to the folder where the files reside: ').strip()

def get_file_names():
    """Get the names of text files to be inserted into the spreadsheet."""
    files = input('Enter the file names you wish to have inserted into the spreadsheet (separated by spaces): ').strip()
    return files.split()

def create_spreadsheet(file_list, path):
    """Create a spreadsheet with the contents of the text files."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    column_num = 1
    for file in file_list:
        file_path = os.path.join(path, file)
        if not os.path.isfile(file_path):
            print(f"File {file} does not exist. Skipping.")
            continue
        
        # Read file contents
        with open(file_path, 'r') as f:
            lines = f.readlines()
        
        # Set the header with bold font
        make_bold = Font(bold=True)
        sheet.cell(row=1, column=column_num, value=file).font = make_bold

        # Write file contents to spreadsheet
        longest = 0
        for row_num, line in enumerate(lines, start=2):
            line = line.strip()
            sheet.cell(row=row_num, column=column_num, value=line)
            longest = max(longest, len(line))
        
        # Adjust column width
        column_letter = get_column_letter(column_num)
        sheet.column_dimensions[column_letter].width = longest
        column_num += 1

    return wb

def main():
    """Main function to execute the script."""
    path = get_input_path()
    file_list = get_file_names()

    if not os.path.isdir(path):
        print(f"The directory {path} does not exist. Please provide a valid path.")
        return

    wb = create_spreadsheet(file_list, path)
    output_file = os.path.join(path, 'text2sheet.xlsx')
    wb.save(output_file)

    print(f"Spreadsheet saved as {output_file} - it can be found in the same directory as the inputted files.")

if __name__ == '__main__':
    main()
