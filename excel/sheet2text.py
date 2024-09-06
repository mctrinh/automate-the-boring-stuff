"""Takes each column of a spreadsheet and saves it to a separate text file."""

import os
import openpyxl

def split_columns_to_text_files(file_path):
    """Splits each column of the spreadsheet into separate text files."""
    
    # Extract path and filename without extension
    path, filename = os.path.split(file_path)
    name, _ = os.path.splitext(filename)
    
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Iterate through each column and write to separate text files
    for col_num in range(1, sheet.max_column + 1):
        col_data = []
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col_num).value
            if cell_value is not None:
                col_data.append(str(cell_value))  # Ensure cell value is string

        # Define the output file path
        output_file_path = os.path.join(path, f'col-{col_num}-{name}.txt')
        
        # Write column data to text file
        with open(output_file_path, 'w') as file:
            file.write('\n'.join(col_data))

        print(f'Column {col_num} saved to {output_file_path}')

def main():
    print('Enter the absolute path of the spreadsheet you wish to split:')
    file_path = input().strip()

    if not os.path.isfile(file_path):
        print(f"Error: The file '{file_path}' does not exist.")
        return

    split_columns_to_text_files(file_path)

if __name__ == '__main__':
    main()
