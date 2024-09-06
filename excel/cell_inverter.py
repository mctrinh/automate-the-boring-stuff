"""Invert the rows and columns of a spreadsheet."""

import re
import openpyxl

def invert_spreadsheet(file_path):
    # Regular expression to split the file path into directory, name, and extension
    path_regex = re.compile(r'(.*/)(.*)(\.xlsx)$')
    path_split = path_regex.search(file_path)
    if not path_split:
        raise ValueError("Invalid file path or not an .xlsx file")

    path = path_split.group(1)
    name = path_split.group(2)
    ext = path_split.group(3)

    # Load the spreadsheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Make a nested list of spreadsheet data (row by row)
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))

    # Open a new spreadsheet and populate it with the inverted data
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    for col_idx, col in enumerate(zip(*rows), start=1):
        for row_idx, value in enumerate(col, start=1):
            new_sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the new spreadsheet with an inverted name
    new_file_path = f"{path}{name}(inverted){ext}"
    new_wb.save(new_file_path)
    print(f"Spreadsheet data inverted and saved to '{new_file_path}'.")

def main():
    print('Enter the absolute path to the spreadsheet file:')
    file_path = input().strip()

    try:
        invert_spreadsheet(file_path)
    except ValueError as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    main()
