"""Creates an N x N multiplication table in a spreadsheet.

Usage: multiplication_table.py <N> - Creates an N x N table
"""

import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def create_multiplication_table(n):
    """Creates an N x N multiplication table and saves it to 'multi_table.xlsx'."""
    
    # Initialize workbook and sheet
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Set up bold font style for headers
    bold_font = Font(bold=True)

    # Write row and column headers
    for i in range(1, n + 1):
        # Set row headers (in the first column)
        sheet.cell(row=i + 1, column=1, value=i).font = bold_font
        # Set column headers (in the first row)
        sheet.cell(row=1, column=i + 1, value=i).font = bold_font

    # Populate the multiplication table
    for row in range(2, n + 2):
        for col in range(2, n + 2):
            cell_formula = f'=A{row}*${get_column_letter(col)}$1'
            sheet.cell(row=row, column=col, value=cell_formula)

    # Save the workbook
    wb.save('multi_table.xlsx')
    print(f"Multiplication table of size {n} x {n} saved to 'multi_table.xlsx'.")

def main():
    if len(sys.argv) != 2:
        print("Usage: multiplication_table.py <N>")
        sys.exit(1)

    try:
        n = int(sys.argv[1])
        if n <= 0:
            raise ValueError
    except ValueError:
        print("Error: <N> must be a positive integer.")
        sys.exit(1)

    create_multiplication_table(n)

if __name__ == '__main__':
    main()
