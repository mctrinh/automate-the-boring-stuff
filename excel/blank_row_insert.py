"""Replicate a spreadsheet in CWD but with N blank rows inserted at a chosen place.

Usage: blank_row_inserter.py takes 3 arguments
<Name> - Name of the spreadsheet to add blank rows to
<Start> - The row to start inserting the blank lines
<Number> - The number of blank lines to insert

e.g. python blank_row_inserter.py test.xlsx 3 2
     Would insert 2 blank rows starting at row 3 into a copy of 'test.xlsx'
"""

import sys
import openpyxl

def insert_blank_rows(filename, start_row, num_blank_rows):
    # Load the existing workbook and sheet
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # Read existing data from the sheet
    rows = list(sheet.iter_rows(values_only=True))
    max_column = len(rows[0]) if rows else 0
    
    # Create a new workbook and sheet
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Insert rows before the blank rows
    print('Inserting blank rows...')
    for row_index, row_data in enumerate(rows):
        if row_index < start_row - 1:
            new_sheet.append(row_data)
        elif row_index == start_row - 1:
            for _ in range(num_blank_rows):
                new_sheet.append([None] * max_column)
            new_sheet.append(row_data)
        else:
            new_sheet.append(row_data)

    # Save the modified workbook
    new_filename = f'blanked-{filename}'
    new_wb.save(new_filename)
    print(f"A copy of the spreadsheet with blanks inserted has been saved as '{new_filename}'. It can be found in the same directory as the original.")

def main():
    if len(sys.argv) != 4:
        print("Usage: python blank_row_inserter.py <Name> <Start> <Number>")
        sys.exit(1)

    name = sys.argv[1]
    blank_start = int(sys.argv[2])
    blank_length = int(sys.argv[3])

    if blank_start < 1 or blank_length < 0:
        print("Start row must be >= 1 and number of blank rows must be >= 0.")
        sys.exit(1)

    insert_blank_rows(name, blank_start, blank_length)

if __name__ == '__main__':
    main()
