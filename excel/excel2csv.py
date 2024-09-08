"""Converts all Excel files in the CWD to similarly named CSV files."""

import os
import csv
import openpyxl

def convert_sheet_to_csv(sheet, csv_writer):
    """Convert a single sheet to CSV format."""
    for row in sheet.iter_rows(values_only=True):
        csv_writer.writerow(row)

def main():
    for file in os.listdir('.'):
        if file.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                csv_file_name = f"{file[:-5]}_{sheet_name}.csv"
                
                with open(csv_file_name, 'w', newline='', encoding='utf-8') as csv_file:
                    csv_writer = csv.writer(csv_file)
                    convert_sheet_to_csv(sheet, csv_writer)

if __name__ == "__main__":
    main()
